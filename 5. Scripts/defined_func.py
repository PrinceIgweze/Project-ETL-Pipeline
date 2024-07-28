from numpy import NaN, clongfloat
import pandas as pd, os, shutil
from datetime import datetime as dt, date as dd, timedelta as td
from dateutil.relativedelta import relativedelta as rd
from parameters import * 
from openpyxl import *



def getReportPeriod():
    cur_month = dd.today() - rd(months= getNoMonths())
    cur_month_actual = dd.today() - rd(months= getNoMonths()+1)
    val = cur_month.strftime("%B_%y")
    val2= cur_month_actual.strftime("%b'%y")
    return [val,val2, cur_month_actual]

def getReportFile():
    
    for xlfiles in os.listdir(getDataSourceFolder()):
        if getReportPeriod()[0] in xlfiles:
            old_path = os.path.join(getDataSourceFolder(),xlfiles)
            new_path = os.path.join(getLandingFolder(), xlfiles)
            shutil.copy(old_path,new_path )

    return new_path    




def getvalues():
    date_time_value = dt.now()
    rp_perd= getReportPeriod()[2]
    destinationWorkBook()
    etl_file_path =os.path.join(getEtlFolder(),"etl_file.xlsx")
    db = load_workbook(etl_file_path)
    ds= db.active
    wb = load_workbook(getReportFile())

    all_data = ({})
    dict_data = {}
    for x in wb.sheetnames:
       
        ws = wb[x]
        #loop through row in m_ws
        for cl in range(1,ws.max_column):
            for rw in range(1,ws.max_row):
                if ws.cell(rw,cl).value is not None:
                    if ws.cell(rw,cl).value == "Item":
                        rpt_per_row = rw 
                        #print(f"Found in row {rpt_per_row} of {x}")
                        break
            
        #this tries to get the col to extract data from
        for col in range(1,ws.max_column):
            if ws.cell(rpt_per_row,col).value is not None:
             
                if getReportPeriod()[1] in ws.cell(rpt_per_row,col).value :
                    col_to_ext = col
                    #print( f"Found in row {rpt_per_row}  column {col} of {x}")
                    break
                    
        for cl in range(1,ws.max_column):
            for rw in range(1,ws.max_row):
                if ws.cell(rw,cl).value is not None:
                    if ws.cell(rw,cl).value == "Cash Paid for":
                        cpf_rw= rw 
                        cpf_col = cl
                        #print( f"Found in row {cpf_rw}  column {cpf_col} of {x}")

                    
        for wr in range(cpf_rw+1, ws.max_row+ 1):
            if ws.cell(wr,cpf_col+1).value is not None:
                prov = ws.title
                item = ws.cell(wr,cpf_col+1).value
                

                nx_row= ds.max_row + 1 
                #for amount
                ds.cell(nx_row  ,1).value = date_time_value
                ds.cell(nx_row  ,2).value = prov
                ds.cell(nx_row  ,3).value = item 
                ds.cell(nx_row  ,4).value = ws.cell(wr,col_to_ext).value
                ds.cell(nx_row  ,5).value = ws.cell(rpt_per_row+1,col_to_ext).value
                ds.cell(nx_row  ,6).value = rp_perd
                #for forecast
                ds.cell(nx_row +1 ,1).value = date_time_value
                ds.cell(nx_row +1 ,2).value = prov
                ds.cell(nx_row +1 ,3).value = item 
                ds.cell(nx_row +1,4).value = ws.cell(wr,col_to_ext+1).value
                ds.cell(nx_row +1,5).value = ws.cell(rpt_per_row+1,col_to_ext+1).value
                ds.cell(nx_row +1,6).value = rp_perd

                
                db.save
    db.save(etl_file_path)
    db.close()            
    return True



def destinationWorkBook():


    wb= Workbook()
    ws= wb.active

    ws.cell(1,1).value= "Date"
    ws.cell(1,2).value= "Province"
    ws.cell(1,3).value= "Cah_Paid_Type"
    ws.cell(1,4).value= "Value"
    ws.cell(1,5).value= "Type"
    ws.cell(1,6).value= "Report_Period"

    wb.save(os.path.join(getEtlFolder(),"etl_file.xlsx"))


print(getvalues())